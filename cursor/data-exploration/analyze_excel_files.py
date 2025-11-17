"""
Excel Data Analysis Script
Analyzes the incentive data Excel files to identify automation opportunities
"""

import pandas as pd
import openpyxl
from pathlib import Path
import json

def analyze_excel_file(file_path):
    """Analyze an Excel file and return detailed information"""
    print(f"\n{'='*80}")
    print(f"Analyzing: {file_path.name}")
    print(f"{'='*80}")
    
    # Load workbook to get sheet information
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    analysis = {
        'filename': file_path.name,
        'sheets': [],
        'total_sheets': len(wb.sheetnames)
    }
    
    for sheet_name in wb.sheetnames:
        print(f"\n[SHEET] {sheet_name}")
        print("-" * 80)
        
        # Read with pandas
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        sheet_info = {
            'name': sheet_name,
            'rows': len(df),
            'columns': len(df.columns),
            'column_names': df.columns.tolist(),
            'dtypes': df.dtypes.astype(str).to_dict(),
            'sample_data': df.head(3).to_dict(orient='records'),
            'missing_values': df.isnull().sum().to_dict(),
            'numeric_columns': df.select_dtypes(include=['number']).columns.tolist(),
            'text_columns': df.select_dtypes(include=['object']).columns.tolist()
        }
        
        print(f"  Dimensions: {len(df)} rows × {len(df.columns)} columns")
        print(f"\n  Columns ({len(df.columns)}):")
        for col in df.columns:
            null_count = df[col].isnull().sum()
            null_pct = (null_count / len(df) * 100) if len(df) > 0 else 0
            print(f"    • {col} ({df[col].dtype}) - {null_count} nulls ({null_pct:.1f}%)")
        
        # Identify potential calculated columns
        if len(df.select_dtypes(include=['number']).columns) > 2:
            print(f"\n  [NUMERIC] Columns: {df.select_dtypes(include=['number']).columns.tolist()}")
        
        # Check for formulas
        ws = wb[sheet_name]
        formula_cells = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_cells.append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })
        
        if formula_cells:
            print(f"\n  [FORMULAS] Found {len(formula_cells)} formula cells:")
            for fc in formula_cells[:5]:  # Show first 5
                print(f"    • {fc['cell']}: {fc['formula']}")
            if len(formula_cells) > 5:
                print(f"    ... and {len(formula_cells) - 5} more")
        
        sheet_info['formula_count'] = len(formula_cells)
        sheet_info['formulas_sample'] = formula_cells[:10]
        
        # Statistical summary for numeric columns
        if len(df.select_dtypes(include=['number']).columns) > 0:
            print(f"\n  [STATS] Statistical Summary:")
            print(df.describe().to_string())
        
        analysis['sheets'].append(sheet_info)
    
    return analysis

def identify_automation_opportunities(analyses):
    """Identify potential automation opportunities based on data analysis"""
    print(f"\n{'='*80}")
    print("[AI] AUTOMATION OPPORTUNITIES IDENTIFIED")
    print(f"{'='*80}")
    
    opportunities = []
    
    for analysis in analyses:
        filename = analysis['filename']
        print(f"\n[FILE] {filename}")
        
        for sheet in analysis['sheets']:
            sheet_name = sheet['name']
            
            # Opportunity 1: Formula automation
            if sheet['formula_count'] > 0:
                opp = f"[+] Replace {sheet['formula_count']} Excel formulas with Python calculations"
                print(f"  * {opp}")
                opportunities.append({
                    'file': filename,
                    'sheet': sheet_name,
                    'type': 'formula_automation',
                    'description': opp,
                    'complexity': 'medium'
                })
            
            # Opportunity 2: Data validation
            if any(v > 0 for v in sheet['missing_values'].values()):
                missing_cols = [k for k, v in sheet['missing_values'].items() if v > 0]
                opp = f"[+] Automated data validation for {len(missing_cols)} columns with missing data"
                print(f"  * {opp}")
                opportunities.append({
                    'file': filename,
                    'sheet': sheet_name,
                    'type': 'data_validation',
                    'description': opp,
                    'complexity': 'low'
                })
            
            # Opportunity 3: Numeric calculations
            if len(sheet['numeric_columns']) > 2:
                opp = f"[+] Automated incentive calculations across {len(sheet['numeric_columns'])} numeric fields"
                print(f"  * {opp}")
                opportunities.append({
                    'file': filename,
                    'sheet': sheet_name,
                    'type': 'calculation_automation',
                    'description': opp,
                    'complexity': 'medium'
                })
            
            # Opportunity 4: Report generation
            if sheet['rows'] > 10:
                opp = f"[+] Automated report generation from {sheet['rows']} records"
                print(f"  * {opp}")
                opportunities.append({
                    'file': filename,
                    'sheet': sheet_name,
                    'type': 'report_generation',
                    'description': opp,
                    'complexity': 'high'
                })
    
    return opportunities

def main():
    # Find Excel files
    test_data_path = Path('test-data')
    excel_files = list(test_data_path.glob('*.xlsx'))
    
    print(f"Found {len(excel_files)} Excel files to analyze")
    
    analyses = []
    for file_path in excel_files:
        analysis = analyze_excel_file(file_path)
        analyses.append(analysis)
    
    # Save detailed analysis
    output_path = Path('cursor/data-exploration/analysis_results.json')
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w') as f:
        json.dump(analyses, f, indent=2, default=str)
    print(f"\n[SAVED] Detailed analysis saved to: {output_path}")
    
    # Identify opportunities
    opportunities = identify_automation_opportunities(analyses)
    
    # Save opportunities
    opp_path = Path('cursor/automation-ideas/identified_opportunities.json')
    opp_path.parent.mkdir(parents=True, exist_ok=True)
    with open(opp_path, 'w') as f:
        json.dump(opportunities, f, indent=2)
    print(f"\n[SAVED] Automation opportunities saved to: {opp_path}")
    
    print(f"\n{'='*80}")
    print(f"[DONE] Analysis Complete! Found {len(opportunities)} automation opportunities")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()

